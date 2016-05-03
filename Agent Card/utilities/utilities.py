import openpyxl
import os

from openpyxl import Workbook
from os import path
from .agent import Agent

SUMMARYPAGE = "Summary"
OUTPUTFILE = "outfile.xlsx"

def openReports():
    while(True): # Attempts to open the raw files
        try:
            agentReport = openpyxl.load_workbook(os.path.dirname(path.dirname(path.abspath(__file__))) + '/raw_files/Agent_Time_Card.xlsx')
            dndReport = openpyxl.load_workbook(os.path.dirname(path.dirname(path.abspath(__file__))) + '/raw_files/Agent_Realtime_Feature_Trace.xlsx')
        except:
            print("**There was an issue with loading one of the agent reports**")
            while(not custom_input("Press enter to try again or CNTL-C to EXIT", "B", "B")):
                pass
            print("Attempting to reload files. Please wait...")
        finally:
            return agentReport, dndReport

def readReports(agentReport, dndReport, timeCard):
    readAgentReport(agentReport, timeCard)
    readDndReport(dndReport, timeCard)

def readAgentReport(agentReport, timeCard):
    agentList = agentReport.get_sheet_by_name(SUMMARYPAGE) # Gets the summary totals for each agent
    for agent in range(5, agentList.max_row + 1):
        currAgent = Agent(valid_input('A', agent, agentList, 'S'))
        currAgent.setTotalTime(valid_input('C', agent, agentList, 'N'))
        timeCard.append(currAgent) # adds agent to the time card with their total login time

def readDndReport(dndReport, timeCard):   # reads the DND report for % availability
    dndList = dndReport.get_sheet_names() # creates a list of all the Agents in the DND rawfile
    dndList.remove(SUMMARYPAGE) # removes the summary page
        
    for index, agent in enumerate(dndList):
        currAgent = dndReport.get_sheet_by_name(agent)
        agentDuration = 0
        for row in range(5, currAgent.max_row): # adds time value of DND for each Agent
            if(valid_input('B', row, currAgent, 'S') == "Do Not Disturb"):
                agentDuration += valid_input('F', row, currAgent, 'N')
        timeCard[index].setDndTime(agentDuration)

def writeTimeCard(timeCard): # opens a new workbook to be the outfile.
    wb = Workbook()          # saves data for each agent to the outfile
    ws = wb.active
    ws['A1'] = "Agent"
    ws['B1'] = "DND Time"
    ws['C1'] = "Total Time"
    ws['D1'] = "% Avail"
    for row, agent in enumerate(timeCard):
        ws['A' + str(row + 2)] = agent.agent
        ws['B' + str(row + 2)] = convert_time_stamp(agent.getDndTime())
        ws['C' + str(row + 2)] = convert_time_stamp(agent.getTotalTime()) 
        ws['D' + str(row + 2)] = agent.getPercentDnd()
    wb.save(os.path.dirname(path.dirname(path.abspath(__file__))) + '/output/' + OUTPUTFILE)
     
def custom_input(input_string, return_type, value_to_find = 'X'): # Prints text and intprets a right/wrong response. Will loop until a correct answer is provided
    while(True):
        if(return_type == 'C'):
            return input(input_string)
        elif(return_type == 'S'):
            return input(input_string)
        elif(return_type == 'B'):
            if(input(input_string)):
                return True
            else:
                return False
        else:
            print("unknown value -> custom_input")
            return
                  
def valid_input(column_position, row, ws, input_type): # Utility function for validating cells from each ws
    if(input_type == 'N'):
        try:
            return int(ws[column_position + str(row)].value)
        except TypeError:
            return 0
        except ValueError:
            return get_sec(ws[column_position + str(row)].value)
    elif(input_type == 'B'):
            return ws[column_position + str(row)].value
    elif(input_type == 'S'):
            return str(ws[column_position + str(row)].value)
    else:
        pass

def get_sec(timeString): # returns time provided as as string in hours, minutes, sec
    try:
        h, m, s = [int(float(i)) for i in timeString.split(':')]
    except:
        return 0
    return convert_sec(h,m,s) # Converts hours, minutes, sec to seconds

def convert_sec(h,m,s): # Converts hours, minutes, sec to seconds
    return (3600 * int(h)) + (60 * int(m)) + int(s)

def convert_time_stamp(seconds): # Converts seconds into a time stamp to be printed in a workbook.
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    return "%d:%02d:%02d" % (h, m, s)
